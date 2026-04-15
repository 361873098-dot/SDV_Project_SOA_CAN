#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DBC to SOA Excel Generator
===========================

Parses a Vector DBC file using `cantools` and generates a structured Excel
workbook that serves as the "service matrix" for M-Core SOA code generation.

The Excel contains the following sheets:
  1. Messages       — CAN message definitions (ID, DLC, direction, cycle time)
  2. Signals        — All signal details (bit layout, factor/offset, type, etc.)
  3. SOA_Service_Map — Pre-filled SOA service mapping template per signal/message
  4. ID_Allocation   — ProviderID / ConsumerID / MethodID allocation table

Usage:
  python dbc_to_soa_excel.py [<input.dbc>] [--node <node_name>] [-o <output.xlsx>]

  If no arguments given, auto-detects DBC from SOA_DBC_File/ subfolder,
  auto-selects the M_CORE node, and outputs to the same folder.

Dependencies:
  pip install cantools openpyxl

Author: Auto-generated tool for SDV SOA project
"""

import os
import sys
import re
import argparse
from pathlib import Path
from datetime import datetime

try:
    import cantools
except ImportError:
    print("ERROR: 'cantools' library is required. Install via: pip install cantools",
          file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: 'openpyxl' library is required. Install via: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# ===========================================================================
#  Constants
# ===========================================================================

# Color palette (hex without #)
COLOR_HEADER_BG   = "1F4E79"   # Dark blue header
COLOR_HEADER_FG   = "FFFFFF"   # White text
COLOR_TX_ROW      = "E2EFDA"   # Light green for TX
COLOR_RX_ROW      = "DAEEF3"   # Light cyan for RX
COLOR_ALT_ROW     = "F2F2F2"   # Light gray for alternating
COLOR_SOA_HEADER  = "7030A0"   # Purple for SOA sheet
COLOR_ID_HEADER   = "C00000"   # Dark red for ID sheet
COLOR_BORDER      = "B4C6E7"   # Light blue border
COLOR_DESC_ROW    = "FFFFCC"   # Light yellow for description row
COLOR_README_BG   = "2E4057"   # Dark teal for README header

# SOA service type options (dropdown candidates for SOA_Service_Map sheet)
# NOTE: Avoid parentheses/slashes in names — Excel DataValidation treats them
#       as formula syntax and corrupts the sheet.
SOA_SERVICE_TYPES = [
    "Event Notifier",
    "Field Notifier",
    "Field Getter",
    "Field Setter",
    "RR Method",
    "Fire-Forget Method",
    "Not Mapped",
]

# IPCF MessageType mapping reference
SOA_TO_IPCF_MAP = {
    "Event Notifier":       "NOTIFICATION_WITHOUT_ACK (0x09)",
    "Field Notifier":       "NOTIFICATION_WITHOUT_ACK (0x09)",
    "Field Getter":         "REQUEST (0x05) / RESPONSE (0x80)",
    "Field Setter":         "REQUEST (0x05) / RESPONSE (0x80)",
    "RR Method":            "REQUEST (0x05) / RESPONSE (0x80)",
    "Fire-Forget Method":   "REQUEST_NO_RETURN_WITHOUT_ACK (0x07)",
    "Not Mapped":           "N/A",
}


# ===========================================================================
#  Style Helpers
# ===========================================================================

def _header_font(bold=True, color=COLOR_HEADER_FG, size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _normal_font(size=10):
    return Font(name="Calibri", size=size)

def _header_fill(color=COLOR_HEADER_BG):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _row_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(top=side, bottom=side, left=side, right=side)

def _center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, row_num, headers, fill_color=COLOR_HEADER_BG):
    """Apply header styling to a row."""
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header_text)
        cell.font = _header_font()
        cell.fill = _header_fill(fill_color)
        cell.alignment = _center_align()
        cell.border = _thin_border()


def _apply_desc_row(ws, row_num, descriptions):
    """Apply a Chinese description sub-header row with light yellow background."""
    for col_idx, desc_text in enumerate(descriptions, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=desc_text)
        cell.font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
        cell.fill = _row_fill(COLOR_DESC_ROW)
        cell.alignment = _center_align()
        cell.border = _thin_border()


def _auto_column_width(ws, min_width=10, max_width=40):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ===========================================================================
#  DBC Helpers
# ===========================================================================

def c_type_for_signal(sig) -> str:
    """Return the best C type string for a DBC signal."""
    if sig.is_float:
        return "float" if sig.length <= 32 else "double"
    if sig.is_signed:
        if sig.length <= 8:    return "int8_t"
        elif sig.length <= 16: return "int16_t"
        elif sig.length <= 32: return "int32_t"
        else:                  return "int64_t"
    else:
        if sig.length <= 8:    return "uint8_t"
        elif sig.length <= 16: return "uint16_t"
        elif sig.length <= 32: return "uint32_t"
        else:                  return "uint64_t"


def upper_name(name: str) -> str:
    """Convert CamelCase / mixed to UPPER_CASE."""
    s1 = re.sub("(.)([A-Z][a-z]+)", r"\1_\2", name)
    result = re.sub("([a-z0-9])([A-Z])", r"\1_\2", s1).upper()
    return re.sub(r"_+", "_", result)


def is_tx(msg, node: str) -> bool:
    if node is None:
        return True
    return node in msg.senders


def is_rx(msg, node: str) -> bool:
    if node is None:
        return True
    for sig in msg.signals:
        if node in sig.receivers:
            return True
    return False


def direction_str(msg, node: str) -> str:
    parts = []
    if is_tx(msg, node):
        parts.append("TX")
    if is_rx(msg, node):
        parts.append("RX")
    return "/".join(parts) if parts else "N/A"


def value_desc_str(sig) -> str:
    """Format value descriptions (choices) into a readable string."""
    if not hasattr(sig, 'choices') or not sig.choices:
        return ""
    items = [f"{v}={d}" for v, d in sorted(sig.choices.items())]
    return "; ".join(items)


def soa_direction_hint(msg, node: str) -> str:
    """Suggest SOA data direction: M->A or A->M."""
    if is_tx(msg, node):
        return "M->A (TX: M-Core sends CAN data to A-Core)"
    elif is_rx(msg, node):
        return "A->M (RX: A-Core controls M-Core to send CAN)"
    return "Unknown"


def default_soa_service(msg, node: str) -> str:
    """Suggest a default SOA service type based on message direction."""
    if is_tx(msg, node):
        # M-Core is sender -> it's providing data -> Event/Notifier
        return "Field Notifier"
    elif is_rx(msg, node):
        # M-Core is receiver -> it accepts control -> Setter
        return "Field Setter"
    return "Not Mapped"


# ===========================================================================
#  Sheet Generators
# ===========================================================================

def create_readme_sheet(wb):
    """Sheet 0: README — overview, field explanations, and usage instructions."""
    ws = wb.active
    ws.title = "README"

    # --- Title ---
    title_cell = ws.cell(row=1, column=1, value="SOA 服务矩阵 Excel 使用说明 (README)")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill = _header_fill(COLOR_README_BG)
    title_cell.alignment = _center_align()
    ws.merge_cells("A1:F1")
    for c in range(2, 7):
        ws.cell(row=1, column=c).fill = _header_fill(COLOR_README_BG)
    ws.row_dimensions[1].height = 35

    # --- Content sections ---
    sections = [
        # (title, content_rows)  where each content_row is (col1_label, col2_desc)
        ("一、工作簿结构 (Sheet Overview)", [
            ("Messages", "CAN 报文定义表。每行一条 CAN 报文，包含 ID、DLC、方向、周期、发送类型等信息。"),
            ("Signals", "CAN 信号定义表。每行一个信号，包含位布局、因子/偏移、数据类型、枚举值等完整信息。"),
            ("SOA_Service_Map", "SOA 服务映射模板（核心编辑页）。将每个 CAN 信号映射为 SOA 服务，定义 IPCF 报文类型和 ID 分配。黄色单元格为用户可编辑字段。"),
            ("ID_Allocation", "ID 分配参考表。展示系统级 ProviderID/ConsumerID 分配规则和每个信号的建议 MethodID。"),
        ]),
        ("二、Messages 表字段说明", [
            ("Send Type (发送类型)",
             "定义 CAN 报文的发送触发策略，来源于 DBC 文件的 GenMsgSendType 属性：\n"
             "  • Cycle — 周期发送：按固定周期持续发送，不论数据是否变化\n"
             "  • Event — 事件触发：仅当信号值变化时发送\n"
             "  • IfActive — 条件激活：满足条件时周期发送，否则停止\n"
             "  • CycleAndEvent — 混合模式：周期发送 + 变化时额外立即发送\n"
             "  • NoMsgSendType — 无发送类型，通常用于被动接收"),
            ("Direction (方向)",
             "以 M-Core 为视角的报文方向：\n"
             "  • TX — M-Core 是发送者，数据从 M→A\n"
             "  • RX — M-Core 是接收者，数据从 A→M\n"
             "  • TX/RX — 该报文同时涉及发送和接收"),
            ("SOA Direction (SOA 数据流)",
             "SOA 架构中的数据流向：\n"
             "  • M→A — M-Core 作为数据提供者，推送 CAN 数据给 A-Core\n"
             "  • A→M — A-Core 作为控制者，发送指令给 M-Core 再转为 CAN 信号"),
        ]),
        ("三、Signals 表字段说明", [
            ("Byte Order (字节序)",
             "定义多字节 CAN 信号在数据帧中的比特排列方式：\n"
             "  • Motorola (BE) — 大端序 / Big-Endian / MSB First：高位字节在前，汽车行业主流\n"
             "  • Intel (LE) — 小端序 / Little-Endian / LSB First：低位字节在前\n"
             "注：IPCF 协议要求大端传输，Motorola 字节序信号无需额外转换，Intel 需翻转"),
            ("Signed (是否有符号)",
             "  • Yes — 有符号数，信号值可以为负数，C 类型为 int8/16/32/64_t\n"
             "  • No — 无符号数，信号值仅为非负数，C 类型为 uint8/16/32/64_t"),
            ("Is Float (是否浮点数)",
             "  • Yes — 信号在 CAN 帧中以 IEEE 754 浮点格式编码（float 32-bit 或 double 64-bit）\n"
             "  • No — 信号以整数编码，物理值 = 原始值 × Factor + Offset\n"
             "注：Factor 不等于 1 并不意味着 Is Float=Yes。Is Float 仅指总线上传输的原始编码方式"),
            ("Factor / Offset (因子 / 偏移量)",
             "物理值与原始值的转换公式：物理值 = 原始值 × Factor + Offset\n"
             "  例：Factor=0.1, Offset=0, 原始值=1500 → 物理值=150.0"),
            ("Value Descriptions (枚举值描述)",
             "信号的离散值与含义的映射关系。例：0=OFF; 1=ON\n"
             "有枚举描述的信号通常表示状态/开关类信号，而非连续数值信号"),
        ]),
        ("四、SOA Service Type 类型详解", [
            ("Event (Notifier)",
             "无状态事件通知。即发即忘，不缓存当前值。\n"
             "  方向：Server → Client（单向）\n"
             "  IPCF：NOTIFICATION_WITHOUT_ACK (0x09)\n"
             "  M核API：PICC_EventSend()\n"
             "  场景：故障上报、一次性状态变化通知"),
            ("Field Notifier",
             "有状态字段变化通知。Server 缓存最新值，值变化时主动推送。\n"
             "  方向：Server → Client（单向）\n"
             "  IPCF：NOTIFICATION_WITHOUT_ACK (0x09)\n"
             "  M核API：PICC_EventSend()\n"
             "  场景：车速、温度等连续数据，M核从CAN接收后推送给A核\n"
             "  与 Event 区别：Field Notifier 缓存当前值，新订阅者可立即获取初始值"),
            ("Field Getter",
             "按需读取字段当前值。Client 发请求，Server 回复当前缓存值。\n"
             "  方向：Client ↔ Server（请求-响应）\n"
             "  IPCF：REQUEST (0x05) / RESPONSE (0x80)\n"
             "  M核API：PICC_GetMethodData() + PICC_MethodResponse()\n"
             "  场景：A核主动读取M核当前CAN信号值"),
            ("Field Setter",
             "远程设置字段新值。Client 发送控制值，Server 执行并回复结果。\n"
             "  方向：Client ↔ Server（请求-响应）\n"
             "  IPCF：REQUEST (0x05) / RESPONSE (0x80)\n"
             "  M核API：PICC_GetMethodData() + PICC_MethodResponse()\n"
             "  场景：A核发控制指令，M核转换为CAN信号发送到总线"),
            ("R/R Method",
             "请求-响应方法调用（Request/Response）。双向通信，异步 SessionID 匹配。\n"
             "  IPCF：REQUEST (0x05) / RESPONSE (0x80)\n"
             "  M核API：取决于角色 — Client 用 PICC_MethodRequest()，Server 用 PICC_GetMethodData()+PICC_MethodResponse()\n"
             "  场景：复杂 RPC 调用，如诊断激活、OTA 命令"),
            ("Fire-and-Forget Method",
             "无需响应的方法调用。发送后立即返回，不等待回复。\n"
             "  IPCF：REQUEST_NO_RETURN_WITHOUT_ACK (0x07)\n"
             "  M核API：PICC_MethodRequest()\n"
             "  场景：不需要确认的简单控制命令"),
            ("Not Mapped",
             "该信号/报文不映射到 SOA 服务，不参与核间通信。"),
        ]),
        ("五、M-Core 约束与注意事项", [
            ("实时性限制",
             "M-Core 不支持同步等待：\n"
             "  • Event 只能发送 WITHOUT_ACK 类型\n"
             "  • Method Request 必须使用异步 SessionID 匹配获取 RESPONSE\n"
             "  • 所有报文均为堆叠报文（10ms 周期或 Buffer 满触发发送）"),
            ("字节序",
             "核间通信协议采用大端 (Big-Endian) 传输。16-bit 以上的数据在打包 Payload 时需注意字节序转换。"),
            ("数据完整性",
             "发送报文时自动附加 2-byte Counter + 2-byte CRC16 校验。Counter 每通道独立累加，断开重连后不重置。"),
            ("连接前置条件",
             "所有 SOA 服务功能必须在 LINK_AVAILABLE 握手成功后才能启动。心跳（Ping/Pong 2秒周期）在两个 IPCF Channel 上双向发送。"),
        ]),
        ("六、使用步骤", [
            ("Step 1", "查看 Messages 和 Signals 表，确认 CAN 报文和信号定义正确。"),
            ("Step 2", "在 SOA_Service_Map 表的黄色单元格中，填写：\n"
                       "  • SOA Service Type — 从下拉菜单选择\n"
                       "  • ProviderID — M核SOA域建议使用 0x71-0x80\n"
                       "  • ConsumerID — A核SOA域建议使用 0x76-0x80\n"
                       "  • MethodID — 同一 ProviderID 下 1-254 唯一"),
            ("Step 3", "参考 ID_Allocation 表确保 ID 不冲突。"),
            ("Step 4", "使用完成后的 Excel 作为输入，运行 C 代码生成工具自动生成 M-Core SOA 代码框架。"),
        ]),
    ]

    row = 3
    for section_title, items in sections:
        # Section title row
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        for label, desc in items:
            cell_l = ws.cell(row=row, column=1, value=label)
            cell_l.font = Font(name="Calibri", bold=True, size=10)
            cell_l.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell_l.border = _thin_border()

            cell_d = ws.cell(row=row, column=2, value=desc)
            cell_d.font = _normal_font()
            cell_d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell_d.border = _thin_border()
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

            # Auto-adjust row height for multi-line content
            line_count = desc.count("\n") + 1
            ws.row_dimensions[row].height = max(15, 15 * line_count)
            row += 1

        row += 1  # blank row between sections

    # Set column widths
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 25

    # Add generation timestamp
    row += 1
    ts_cell = ws.cell(row=row, column=1,
                      value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ts_cell.font = Font(name="Calibri", size=9, italic=True, color="999999")

    return ws


def create_messages_sheet(wb, db, node):
    """Sheet 1: Messages — one row per CAN message."""
    ws = wb.create_sheet(title="Messages")

    headers = [
        "No.",
        "Message Name",
        "CAN ID (Hex)",
        "CAN ID (Dec)",
        "DLC (Bytes)",
        "Direction (M-Core)",
        "Sender Node",
        "Cycle Time (ms)",
        "Send Type",
        "Signal Count",
        "SOA Direction",
        "Comment",
    ]
    descriptions = [
        "序号",
        "CAN报文名称",
        "CAN ID（十六进制）",
        "CAN ID（十进制）",
        "数据长度（字节数）",
        "M核视角: TX=M核发送, RX=M核接收",
        "DBC中定义的发送节点",
        "周期发送间隔(ms), 空=非周期",
        "发送触发类型: Cycle=周期, Event=事件触发, IfActive=条件激活",
        "该报文包含的信号个数",
        "SOA数据流方向: M→A=M核推送, A→M=A核控制",
        "DBC中的报文注释",
    ]
    _apply_header_row(ws, 1, headers)
    _apply_desc_row(ws, 2, descriptions)

    # Freeze top 2 rows (header + description)
    ws.freeze_panes = "A3"

    row_num = 3
    for idx, msg in enumerate(sorted(db.messages, key=lambda m: m.frame_id), start=1):
        msg_is_tx = is_tx(msg, node)
        msg_is_rx = is_rx(msg, node)
        if not (msg_is_tx or msg_is_rx):
            continue

        dir_tag = direction_str(msg, node)
        sender = ", ".join(msg.senders) if msg.senders else "N/A"
        cycle = msg.cycle_time if msg.cycle_time else ""
        send_type = ""
        # Try to get GenMsgSendType from DBC attributes
        try:
            send_type = msg.dbc.attributes.get("GenMsgSendType", {}).value
        except (AttributeError, KeyError):
            pass
        comment = msg.comment if msg.comment else ""

        values = [
            idx,
            msg.name,
            f"0x{msg.frame_id:03X}",
            msg.frame_id,
            msg.length,
            dir_tag,
            sender,
            cycle,
            send_type,
            len(msg.signals),
            soa_direction_hint(msg, node),
            comment,
        ]

        fill_color = COLOR_TX_ROW if msg_is_tx else COLOR_RX_ROW
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = _normal_font()
            cell.fill = _row_fill(fill_color)
            cell.alignment = _center_align() if col_idx <= 10 else _left_align()
            cell.border = _thin_border()

        row_num += 1

    _auto_column_width(ws)
    return ws


def create_signals_sheet(wb, db, node):
    """Sheet 2: Signals — one row per signal with full details."""
    ws = wb.create_sheet(title="Signals")

    headers = [
        "No.",
        "Message Name",
        "CAN ID (Hex)",
        "Signal Name",
        "C Macro Name",
        "Start Bit",
        "Bit Length",
        "Byte Order",
        "Signed",
        "Is Float",
        "C Data Type",
        "Factor",
        "Offset",
        "Min",
        "Max",
        "Unit",
        "Value Descriptions",
        "Init Value",
        "Receiver Nodes",
        "Comment",
    ]
    descriptions = [
        "序号",
        "所属报文名称",
        "所属报文CAN ID",
        "信号名称（DBC原始名）",
        "C代码中的宏定义名(自动生成)",
        "信号起始位（DBC定义）",
        "信号位宽度",
        "字节序: Motorola(BE)=大端/高位在前, Intel(LE)=小端/低位在前",
        "是否有符号: Yes=可为负数(int), No=仅非负(uint)",
        "是否浮点: Yes=IEEE754浮点编码, No=整数编码",
        "映射的C数据类型（基于位宽/符号/浮点自动推导）",
        "物理值=原始值×Factor+Offset中的Factor",
        "物理值=原始值×Factor+Offset中的Offset",
        "物理值最小值",
        "物理值最大值",
        "物理值单位",
        "枚举值描述（如 0=OFF; 1=ON）",
        "信号初始值（DBC GenSigStartValue属性）",
        "DBC中定义的接收节点",
        "DBC中的信号注释",
    ]
    _apply_header_row(ws, 1, headers)
    _apply_desc_row(ws, 2, descriptions)
    ws.freeze_panes = "A3"

    row_num = 3
    sig_no = 1
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        msg_is_tx = is_tx(msg, node)
        msg_is_rx = is_rx(msg, node)
        if not (msg_is_tx or msg_is_rx):
            continue

        for sig in msg.signals:
            byte_order = "Motorola (BE)" if sig.byte_order == "big_endian" else "Intel (LE)"
            receivers = ", ".join(sig.receivers) if sig.receivers else "N/A"
            init_val = ""
            try:
                init_val = sig.dbc.attributes.get("GenSigStartValue", {}).value
            except (AttributeError, KeyError):
                pass
            comment = sig.comment if sig.comment else ""

            values = [
                sig_no,
                msg.name,
                f"0x{msg.frame_id:03X}",
                sig.name,
                upper_name(sig.name),
                sig.start,
                sig.length,
                byte_order,
                "Yes" if sig.is_signed else "No",
                "Yes" if sig.is_float else "No",
                c_type_for_signal(sig),
                sig.scale,
                sig.offset,
                sig.minimum,
                sig.maximum,
                sig.unit if sig.unit else "",
                value_desc_str(sig),
                init_val,
                receivers,
                comment,
            ]

            fill_color = COLOR_ALT_ROW if sig_no % 2 == 0 else "FFFFFF"
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = _normal_font()
                if fill_color != "FFFFFF":
                    cell.fill = _row_fill(fill_color)
                cell.alignment = _center_align() if col_idx <= 15 else _left_align()
                cell.border = _thin_border()

            row_num += 1
            sig_no += 1

    _auto_column_width(ws)
    return ws


def create_soa_service_map_sheet(wb, db, node):
    """Sheet 3: SOA_Service_Map — editable template for SOA service assignment."""
    ws = wb.create_sheet(title="SOA_Service_Map")

    headers = [
        "No.",
        "Message Name",
        "CAN ID (Hex)",
        "Signal Name",
        "CAN Direction",
        "SOA Data Flow",
        "SOA Service Type",
        "IPCF MessageType Mapping",
        "M-Core PICC API",
        "ProviderID (Hex)",
        "ConsumerID (Hex)",
        "MethodID (Dec)",
        "SOA Service Name",
        "Payload Struct Name",
        "Trigger Condition",
        "Period (ms)",
        "Description",
    ]
    descriptions = [
        "序号",
        "CAN报文名称（来源DBC）",
        "CAN ID（十六进制）",
        "CAN信号名称",
        "CAN方向: TX=M核发送, RX=M核接收",
        "SOA数据流: M→A或A→M",
        "SOA服务类型（下拉选择）",
        "对应IPCF私有协议报文类型（自动映射）",
        "M核使用的PICC API函数",
        "服务提供者ID(0x01-0xFE)",
        "服务消费者ID(0x01-0xFE)",
        "方法ID(1-254, 同一Provider下唯一)",
        "SOA服务名称（自动生成,可修改）",
        "Payload结构体名称（用于代码生成）",
        "发送触发条件",
        "发送周期(ms)",
        "功能描述说明",
    ]
    _apply_header_row(ws, 1, headers, fill_color=COLOR_SOA_HEADER)
    _apply_desc_row(ws, 2, descriptions)
    ws.freeze_panes = "A3"

    # IPCF -> M-Core API mapping
    soa_to_api_map = {
        "Event Notifier":       "PICC_EventSend()",
        "Field Notifier":       "PICC_EventSend()",
        "Field Getter":         "PICC_GetMethodData() / PICC_MethodResponse()",
        "Field Setter":         "PICC_GetMethodData() / PICC_MethodResponse()",
        "RR Method":            "PICC_MethodRequest() or PICC_GetMethodData()/PICC_MethodResponse()",
        "Fire-Forget Method":   "PICC_MethodRequest()",
        "Not Mapped":           "N/A",
    }

    row_num = 3
    entry_no = 1
    data_start_row = 3  # remember first data row for data validation

    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        msg_is_tx = is_tx(msg, node)
        msg_is_rx = is_rx(msg, node)
        if not (msg_is_tx or msg_is_rx):
            continue

        dir_tag = direction_str(msg, node)
        soa_flow = soa_direction_hint(msg, node)
        soa_svc = default_soa_service(msg, node)
        ipcf_map = SOA_TO_IPCF_MAP.get(soa_svc, "N/A")
        api_map = soa_to_api_map.get(soa_svc, "N/A")

        # Default IDs (to be filled by user / tool)
        provider_id = ""
        consumer_id = ""
        method_id = ""

        # Suggest trigger condition based on direction
        if msg_is_tx:
            trigger = "On Change / Cyclic"
        else:
            trigger = "On Request (from A-Core)"

        cycle = msg.cycle_time if msg.cycle_time else ""

        # Create one row per signal for fine-grained SOA mapping
        for sig in msg.signals:
            # Auto-generate suggested SOA service name and struct name
            soa_name = f"Soa_{msg.name}_{sig.name}"
            struct_name = f"{msg.name}_{sig.name}_Payload_t"

            values = [
                entry_no,
                msg.name,
                f"0x{msg.frame_id:03X}",
                sig.name,
                dir_tag,
                soa_flow,
                soa_svc,
                ipcf_map,
                api_map,
                provider_id,
                consumer_id,
                method_id,
                soa_name,
                struct_name,
                trigger,
                cycle,
                f"CAN signal '{sig.name}' mapped to SOA service",
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = _normal_font()
                cell.alignment = _center_align() if col_idx <= 12 else _left_align()
                cell.border = _thin_border()
                # Highlight editable columns with light yellow
                if col_idx in (7, 10, 11, 12, 13, 14, 15, 16):
                    cell.fill = _row_fill("FFF2CC")  # Light yellow = user editable

            row_num += 1
            entry_no += 1

    # Add legend / note rows
    row_num += 1
    legend_items = [
        ("■  图例说明 (Legend)  ■", ""),
        ("黄色单元格", "用户可编辑字段。请填入 ProviderID、ConsumerID、MethodID，并根据需要调整 SOA Service Type。"),
        ("", ""),
        ("■  SOA Service Type 类型说明  ■", ""),
        ("Event Notifier",
         "无状态事件通知。单向: Server→Client。即发即忘，不缓存当前值。\n"
         "IPCF: NOTIFICATION_WITHOUT_ACK (0x09) | M核API: PICC_EventSend()"),
        ("Field Notifier",
         "有状态字段变化通知。Server缓存最新值，值变化时主动推送。新订阅者可获取初始值。\n"
         "IPCF: NOTIFICATION_WITHOUT_ACK (0x09) | M核API: PICC_EventSend()\n"
         "典型场景: 车速、温度等连续数据，M核CAN接收后推送给A核"),
        ("Field Getter",
         "按需读取字段当前值。Client发请求→Server回复当前缓存值。\n"
         "IPCF: REQUEST(0x05)/RESPONSE(0x80) | M核API: PICC_GetMethodData()+PICC_MethodResponse()\n"
         "典型场景: A核读取M核CAN信号当前值"),
        ("Field Setter",
         "远程设置字段新值。Client发送控制值→Server执行并回复结果。\n"
         "IPCF: REQUEST(0x05)/RESPONSE(0x80) | M核API: PICC_GetMethodData()+PICC_MethodResponse()\n"
         "典型场景: A核发控制指令→M核转换为CAN信号发送到总线"),
        ("RR Method",
         "请求-响应方法调用（Request/Response）。双向通信，异步SessionID匹配。\n"
         "IPCF: REQUEST(0x05)/RESPONSE(0x80) | M核API: 取决于角色(Client/Server)\n"
         "典型场景: 复杂RPC调用，如诊断激活、OTA命令"),
        ("Fire-Forget Method",
         "无需响应的方法调用。发送后立即返回，不等待任何回复。\n"
         "IPCF: REQUEST_NO_RETURN_WITHOUT_ACK (0x07) | M核API: PICC_MethodRequest()"),
        ("Not Mapped",
         "该信号/报文不映射到SOA服务，不参与核间通信。"),
        ("", ""),
        ("■  ID 分配规则  ■", ""),
        ("ProviderID/ConsumerID",
         "范围: 0x01-0xFE。系统全局唯一。一个Provider可服务多个Consumer，一个Consumer只能关联一个Provider。\n"
         "M核SOA域: 0x71-0x80 | A核SOA域: 0x76-0x80"),
        ("MethodID",
         "范围: 1-254。在同一个ProviderID下唯一。Event ID和Method ID不可重叠。"),
        ("", ""),
        ("■  M核约束  ■", ""),
        ("M核实时性限制",
         "M核不支持同步等待。Event只能发WITHOUT_ACK类型。Method Request必须使用异步SessionID匹配获取RESPONSE。\n"
         "所有报文均为堆叠报文(10ms周期或Buffer满触发发送)。数据传输采用大端字节序。"),
    ]
    for label, desc in legend_items:
        cell_l = ws.cell(row=row_num, column=1, value=label)
        if label.startswith("\u25a0"):
            cell_l.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        else:
            cell_l.font = Font(name="Calibri", bold=True, size=10, color="7030A0")
        cell_d = ws.cell(row=row_num, column=2, value=desc)
        cell_d.font = _normal_font()
        cell_d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=17)
        # Set row height for multi-line descriptions
        if "\n" in desc:
            ws.row_dimensions[row_num].height = 15 * (desc.count("\n") + 1)
        row_num += 1

    _auto_column_width(ws)
    return ws


def create_id_allocation_sheet(wb, db, node):
    """Sheet 4: ID_Allocation — ProviderID / ConsumerID allocation reference."""
    ws = wb.create_sheet(title="ID_Allocation")

    # Part 1: System ID Allocation Table (from IPCF spec)
    headers_1 = [
        "ID Range",
        "Function Domain",
        "M-Core IDs",
        "A-Core IDs",
        "Status",
    ]
    _apply_header_row(ws, 1, headers_1, fill_color=COLOR_ID_HEADER)

    id_ranges = [
        ("01 - 10", "Power Management",   "0x01", "0x06", "Reserved"),
        ("11 - 20", "OTA",                "0x11", "0x16", "Reserved"),
        ("21 - 30", "Health Management",   "0x21", "0x26", "Reserved"),
        ("31 - 40", "Communication Module","0x31", "0x36", "Reserved"),
        ("41 - 50", "Storage Module",      "0x41", "0x46", "Reserved"),
        ("51 - 60", "Diagnostic Module",   "0x51", "0x56", "Reserved"),
        ("61 - 70", "Time Sync",           "0x61", "0x66", "Reserved"),
        ("71 - 80", "SOA",                 "0x71", "0x76", "For SOA CAN Services"),
        ("81 - 254","Reserved",            "--",   "--",   "Available"),
    ]

    row_num = 2
    for rng, func, m_id, a_id, status in id_ranges:
        values = [rng, func, m_id, a_id, status]
        fill = COLOR_TX_ROW if status == "For SOA CAN Services" else "FFFFFF"
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = _normal_font()
            if fill != "FFFFFF":
                cell.fill = _row_fill(fill)
            cell.alignment = _center_align()
            cell.border = _thin_border()
        row_num += 1

    # Part 2: Per-message MethodID allocation suggestion
    row_num += 2
    ws.cell(row=row_num, column=1, value="Per-Message MethodID Allocation").font = \
        Font(name="Calibri", bold=True, size=12, color=COLOR_ID_HEADER)
    row_num += 1

    headers_2 = [
        "Message Name",
        "CAN ID (Hex)",
        "Direction",
        "Suggested ProviderID",
        "Suggested ConsumerID",
        "Signal Name",
        "Suggested MethodID",
        "SOA Service Type",
        "Notes",
    ]
    _apply_header_row(ws, row_num, headers_2, fill_color=COLOR_ID_HEADER)
    row_num += 1

    method_id_counter = 1
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        msg_is_tx = is_tx(msg, node)
        msg_is_rx = is_rx(msg, node)
        if not (msg_is_tx or msg_is_rx):
            continue

        dir_tag = direction_str(msg, node)
        soa_svc = default_soa_service(msg, node)

        for sig in msg.signals:
            # Auto-suggest IDs from SOA range
            if msg_is_tx:
                # M-Core provides data -> M is Provider
                prov_id = "0x71 (M-Core)"
                cons_id = "0x76 (A-Core)"
            else:
                # A-Core sends commands -> A is Provider, M is Consumer
                prov_id = "0x76 (A-Core)"
                cons_id = "0x71 (M-Core)"

            notes = "Auto-suggested, adjust as needed"

            values = [
                msg.name,
                f"0x{msg.frame_id:03X}",
                dir_tag,
                prov_id,
                cons_id,
                sig.name,
                method_id_counter,
                soa_svc,
                notes,
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = _normal_font()
                cell.alignment = _center_align() if col_idx <= 7 else _left_align()
                cell.border = _thin_border()

            row_num += 1
            method_id_counter += 1

    _auto_column_width(ws)
    return ws


# ===========================================================================
#  Auto-detection helpers
# ===========================================================================

def _auto_find_dbc(script_dir: Path) -> Path:
    """Search for .dbc files in the SOA_DBC_File subfolder next to this script."""
    dbc_dir = script_dir / "SOA_DBC_File"
    if not dbc_dir.is_dir():
        print(f"ERROR: DBC folder not found: {dbc_dir}", file=sys.stderr)
        sys.exit(1)
    dbc_files = list(dbc_dir.glob("*.dbc"))
    if len(dbc_files) == 0:
        print(f"ERROR: No .dbc files found in {dbc_dir}", file=sys.stderr)
        sys.exit(1)
    if len(dbc_files) > 1:
        print(f"WARNING: Multiple .dbc files found, using the first one: {dbc_files[0].name}")
    return dbc_files[0]


def _auto_select_node(db) -> str:
    """Auto-select the M-Core node from the DBC database."""
    node_names = [n.name for n in db.nodes]
    for name in node_names:
        if "M_CORE" in name.upper():
            return name
    for name in node_names:
        upper = name.upper()
        if "MCORE" in upper or "MCU" in upper:
            return name
    print("WARNING: No M-Core node found, generating for ALL nodes.")
    return None


# ===========================================================================
#  Main
# ===========================================================================

def main():
    ap = argparse.ArgumentParser(
        description="Parse DBC file and generate a structured SOA Excel workbook.\n"
                    "When run WITHOUT arguments, auto-detects DBC from SOA_DBC_File/,\n"
                    "auto-selects M_CORE node, outputs Excel to the same folder.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("dbc", nargs="?", default=None,
                    help="Path to the input .dbc file (auto-detected if omitted)")
    ap.add_argument("--node", default=None,
                    help="Target ECU node name (auto-detected if omitted)")
    ap.add_argument("-o", "--output", default=None,
                    help="Output Excel file path (auto-generated if omitted)")
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent

    # --- Auto-detect DBC file ---
    if args.dbc is not None:
        dbc_path = Path(args.dbc).resolve()
    else:
        dbc_path = _auto_find_dbc(script_dir)
        print(f"[Auto] DBC file: {dbc_path}")

    if not dbc_path.is_file():
        print(f"ERROR: DBC file not found: {dbc_path}", file=sys.stderr)
        sys.exit(1)

    # --- Parse DBC ---
    db = cantools.database.load_file(str(dbc_path))
    print(f"[cantools] Parsed {len(db.messages)} messages, "
          f"nodes: {[n.name for n in db.nodes]}")

    # --- Auto-detect node ---
    if args.node is not None:
        node = args.node
    else:
        node = _auto_select_node(db)
        if node:
            print(f"[Auto] Selected node: {node}")

    # --- Output path ---
    if args.output is not None:
        out_path = Path(args.output).resolve()
    else:
        out_path = dbc_path.parent / f"{dbc_path.stem}_SOA_Matrix.xlsx"
        print(f"[Auto] Output: {out_path}")

    # --- Create Workbook ---
    wb = Workbook()

    print("[*] Creating 'README' sheet...")
    create_readme_sheet(wb)

    print("[*] Creating 'Messages' sheet...")
    create_messages_sheet(wb, db, node)

    print("[*] Creating 'Signals' sheet...")
    create_signals_sheet(wb, db, node)

    print("[*] Creating 'SOA_Service_Map' sheet...")
    create_soa_service_map_sheet(wb, db, node)

    print("[*] Creating 'ID_Allocation' sheet...")
    create_id_allocation_sheet(wb, db, node)

    # --- Save ---
    wb.save(str(out_path))
    print(f"\n{'='*60}")
    print(f"  SUCCESS: Excel saved to:")
    print(f"  {out_path}")
    print(f"{'='*60}")

    # --- Summary ---
    print(f"\nSummary:")
    total_signals = 0
    for msg in sorted(db.messages, key=lambda m: m.frame_id):
        msg_is_tx = is_tx(msg, node)
        msg_is_rx = is_rx(msg, node)
        if not (msg_is_tx or msg_is_rx):
            continue
        dir_tag = direction_str(msg, node)
        soa_svc = default_soa_service(msg, node)
        print(f"  0x{msg.frame_id:03X}  {msg.name:30s}  DLC={msg.length}  "
              f"[{dir_tag}]  signals={len(msg.signals)}  SOA={soa_svc}")
        total_signals += len(msg.signals)

    print(f"\n  Total messages: {len(db.messages)}")
    print(f"  Total signals:  {total_signals}")
    print(f"  Target node:    {node or 'ALL'}")
    print(f"\nNext step: Edit the 'SOA_Service_Map' sheet (yellow cells) to assign")
    print(f"ProviderID, ConsumerID, MethodID, then use the Excel to generate C code.")


if __name__ == "__main__":
    main()
