#!C:/Users/RogPlus/AppData/Local/Programs/Python/Python312/python.exe
# -*- coding: utf-8 -*-
"""
SOA Adapter Code Generator
===========================

Reads a Vector DBC file and an SOA service matrix Excel workbook,
matches signals between them, and auto-generates SOA_Adapter_Generated.c/.h
that replace the hand-maintained soa_adapter_cnf.c/.h.

Signal Interface Convention:
  - CAN RX signals (sender != M_CORE, real message)   → Can_Get_Rx_signal_<Name>()
  - CAN TX signals (sender == M_CORE, real message)   → Can_R_W_signal_<Name>()
  - Application-layer signals (virtual msg / no msg)  → APP_SW_R_W_signal_<Name>()

Usage:
    python soa_adapter_generator.py [--dbc <input.dbc>] [--excel <input.xlsx>] [-o <output_dir>]

    If no arguments given, auto-detects inputs from SOA_DBC_File/ subfolder.

Dependencies:
    pip install cantools openpyxl
"""

import os
import sys
import re
import argparse
import math
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple

# =====================================================================
#  Path Configuration
# =====================================================================
BASE_PROJECT_DIR = Path(
    r"D:\MySandbox\SDV_Project\IPCF_FreeRTOS_S32G399A_M7_Oring"
)

DEFAULT_DBC_FILE_PATH = (
    BASE_PROJECT_DIR / "Tools" / "CAN_Tools" / "SOA_DBC_File" / "CANdbc_file.dbc"
)

DEFAULT_EXCEL_FILE_PATH = (
    BASE_PROJECT_DIR / "Tools" / "CAN_Tools" / "SOA_DBC_File" / "S2S_HPC_A_SAG.xlsx"
)

DEFAULT_OUTPUT_DIR = (
    BASE_PROJECT_DIR / "Tools" / "CAN_Tools" / "SOA_DBC_File"
)

DEFAULT_OUTPUT_BASENAME = "SOA_Adapter_Generated"

# Virtual message name in DBC (Vector placeholder for orphan signals)
VIRTUAL_MSG_NAME = "VECTOR__INDEPENDENT_SIG_MSG"

# =====================================================================
#  Dependency check
# =====================================================================
try:
    import cantools
except ImportError:
    print("ERROR: 'cantools' library is required. Install via: pip install cantools",
          file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: 'openpyxl' library is required. Install via: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


# =====================================================================
#  Data Models
# =====================================================================

@dataclass
class DBCSignalInfo:
    """Metadata for a single DBC signal, resolved from cantools parsing."""
    signal_name: str           # DBC signal name (e.g. "VehicleSpeed")
    message_name: str          # Parent message name (e.g. "Standard_200_Rx")
    message_id: int           # CAN ID
    sender: str                # Message sender node name
    is_tx: bool                # True if sender == M_CORE node
    is_virtual: bool           # True if parent message is VECTOR__INDEPENDENT_SIG_MSG
    c_type: str               # C type string (uint8_t, uint16_t, etc.)
    byte_size: int             # Serialized byte count (1/2/4/8)
    bit_length: int            # Signal bit width
    factor: float              # Signal factor
    offset: float              # Signal offset
    unit: str                  # Signal unit
    min_val: float             # Minimum value
    max_val: float             # Maximum value
    is_signed: bool            # Whether signal is signed
    is_float: bool             # Whether signal is floating-point

    @property
    def is_app_signal(self) -> bool:
        """True if this signal lives in a virtual message → application-layer signal."""
        return self.is_virtual

    @property
    def is_can_rx(self) -> bool:
        """True if this is a real CAN RX signal (not TX, not virtual)."""
        return (not self.is_tx) and (not self.is_virtual)

    @property
    def is_can_tx(self) -> bool:
        """True if this is a real CAN TX signal (TX, not virtual)."""
        return self.is_tx and (not self.is_virtual)

    @property
    def can_access_func(self) -> str:
        """Return the C function name to access this signal."""
        if self.is_can_rx:
            return f"Can_Get_Rx_signal_{self.signal_name}"
        elif self.is_can_tx:
            return f"Can_R_W_signal_{self.signal_name}"
        else:
            return f"APP_SW_R_W_signal_{self.signal_name}"


@dataclass
class SOAServiceEntry:
    """One row in the SOA service table, fully resolved after matching."""
    # --- Identity (from Excel) ---
    service_interface_name: str   # e.g. "Atom_VCU_DriSpeedSt"
    service_id: int               # e.g. 0x0001
    service_description: str      # e.g. "上报当前车速"
    service_type: str             # "Notifier" / "Getter" / "Setter"
    method_id: int                 # e.g. 0x8001
    eventgroup_name: str          # e.g. "EG_DriSpeedSwtSts" or ""
    eventgroup_id: int            # e.g. 0x0001 or 0
    # --- Signal matching (from DBC) ---
    signal_info: Optional[DBCSignalInfo] = None
    # --- Derived (computed) ---
    table_index: int = 0
    has_linked_notifier: bool = False
    linked_notifier_idx: int = 0

    # --- Convenience properties ---

    @property
    def soa_service_type_enum(self) -> str:
        mapping = {"Notifier": "SOA_SERVICE_NOTIFIER", "Getter": "SOA_SERVICE_GETTER", "Setter": "SOA_SERVICE_SETTER"}
        return mapping.get(self.service_type, "SOA_SERVICE_NOTIFIER")

    @property
    def soa_service_type_c(self) -> str:
        return self.soa_service_type_enum + ","

    @property
    def is_notifier(self) -> bool:
        return self.service_type == "Notifier"

    @property
    def is_getter(self) -> bool:
        return self.service_type == "Getter"

    @property
    def is_setter(self) -> bool:
        return self.service_type == "Setter"

    @property
    def has_read_func(self) -> bool:
        return self.is_notifier or self.is_getter

    @property
    def has_write_func(self) -> bool:
        return self.is_setter

    @property
    def data_size(self) -> int:
        if self.signal_info:
            return self.signal_info.byte_size
        return 1  # default 1 byte if unknown

    @property
    def module_prefix(self) -> str:
        """Extract module prefix from service interface name: Atom_VCU_xxx -> VCU, Atom_BCM_xxx -> BCM."""
        parts = self.service_interface_name.split("_")
        if len(parts) >= 3:
            return parts[1]  # VCU, BCM, etc.
        return "GEN"

    @property
    def short_name(self) -> str:
        """Extract the descriptive part after module prefix: Atom_VCU_DriSpeedSt -> DriSpeed."""
        parts = self.service_interface_name.split("_")
        if len(parts) >= 3:
            # Join everything after module prefix, then strip common suffixes
            name = "_".join(parts[2:])
            # Remove common suffixes like "St"
            for suffix in ["St"]:
                if name.endswith(suffix) and len(name) > len(suffix) + 2:
                    name = name[:-len(suffix)]
            return name
        return parts[-1] if parts else "Unknown"

    @property
    def read_func_name(self) -> str:
        if self.signal_info:
            return f"SOA_Read{self.signal_info.signal_name}"
        return f"SOA_Read{self.short_name}"

    @property
    def write_func_name(self) -> str:
        if self.signal_info:
            return f"SOA_Write{self.signal_info.signal_name}"
        return f"SOA_Write{self.short_name}"

    @property
    def idx_macro_name(self) -> str:
        """e.g. SOA_SVC_IDX_DRI_SPEED_NOTIF"""
        type_suffix = {"Notifier": "NOTIF", "Getter": "GETTER", "Setter": "SETTER"}
        return f"SOA_SVC_IDX_{upper_name(self.short_name)}_{type_suffix.get(self.service_type, 'UNKNOWN')}"

    @property
    def sid_macro_name(self) -> str:
        """e.g. SOA_SID_VCU_DRI_SPEED"""
        return f"SOA_SID_{self.module_prefix}_{upper_name(self.short_name)}"

    @property
    def mid_macro_name(self) -> str:
        """e.g. SOA_MID_DRI_SPEED_NOTIF"""
        type_suffix = {"Notifier": "NOTIF", "Getter": "GETTER", "Setter": "SETTER"}
        return f"SOA_MID_{upper_name(self.short_name)}_{type_suffix.get(self.service_type, 'UNKNOWN')}"

    @property
    def eg_macro_name(self) -> str:
        """e.g. SOA_EG_DRI_SPEED_SWTS"""
        if self.eventgroup_name:
            # Extract meaningful part: EG_DriSpeedSwtSts -> DRI_SPEED_SWTS
            eg_part = self.eventgroup_name
            if eg_part.startswith("EG_"):
                eg_part = eg_part[3:]
            return f"SOA_EG_{upper_name(eg_part)}"
        return ""


# =====================================================================
#  Helpers
# =====================================================================

def upper_name(name: str) -> str:
    """Convert CamelCase / mixed to UPPER_CASE, collapsing double underscores."""
    s1 = re.sub("(.)([A-Z][a-z]+)", r"\1_\2", name)
    result = re.sub("([a-z0-9])([A-Z])", r"\1_\2", s1).upper()
    result = re.sub(r"_+", "_", result)
    return result


def c_type_for_signal(sig) -> str:
    """Return the best C type for a signal's struct member."""
    if sig.is_float:
        if sig.length <= 32:
            return "float"
        else:
            return "double"
    if sig.is_signed:
        if sig.length <= 8:
            return "int8_t"
        elif sig.length <= 16:
            return "int16_t"
        elif sig.length <= 32:
            return "int32_t"
        else:
            return "int64_t"
    else:
        if sig.length <= 8:
            return "uint8_t"
        elif sig.length <= 16:
            return "uint16_t"
        elif sig.length <= 32:
            return "uint32_t"
        else:
            return "uint64_t"


def byte_size_for_signal(sig) -> int:
    """Return the serialized byte size for a signal."""
    bits = sig.length
    if bits <= 8:
        return 1
    elif bits <= 16:
        return 2
    elif bits <= 32:
        return 4
    else:
        return 8


def c_type_for_byte_size(byte_size: int) -> str:
    """Map byte size back to a C unsigned type for local variable."""
    mapping = {1: "uint8_t", 2: "uint16_t", 4: "uint32_t", 8: "uint64_t"}
    return mapping.get(byte_size, "uint8_t")


def format_hex16(val: int) -> str:
    """Format a 16-bit value as C hex literal: 0x0001U"""
    return f"0x{val:04X}U"


def format_hex8(val: int) -> str:
    """Format a 8-bit value as C hex literal: 0x01U"""
    return f"0x{val:02X}U"


# =====================================================================
#  DBC Parsing
# =====================================================================

def _parse_dbc_virtual_signals(dbc_path: Path, node_name: str) -> Dict[str, DBCSignalInfo]:
    """
    Manually parse VECTOR__INDEPENDENT_SIG_MSG signals from the DBC file.

    cantools often skips virtual messages (DLC=0). This function uses regex
    to extract signal definitions from the raw DBC text.
    """
    signal_map: Dict[str, DBCSignalInfo] = {}
    dbc_text = dbc_path.read_text(encoding="utf-8", errors="replace")

    # Pattern: BO_ <id> VECTOR__INDEPENDENT_SIG_MSG: <dlc> <sender>
    #          SG_ <name> : <start>|<length>@<endian><sign> (<factor>,<offset>) [<min>|<max>] "<unit>" <receiver>
    bo_pattern = re.compile(
        r"BO_\s+(\d+)\s+" + re.escape(VIRTUAL_MSG_NAME) + r"\s*:\s*\d+\s+(\S+)"
    )
    sg_pattern = re.compile(
        r"SG_\s+(\S+)\s*:\s*(\d+)\|(\d+)@(\d)([+-])\s*"
        r"\(([^,]+),([^)]+)\)\s*"
        r"\[([^|]+)\|([^\]]+)\]\s*"
        r'"([^"]*)"'
    )

    bo_match = bo_pattern.search(dbc_text)
    if not bo_match:
        return signal_map  # No virtual message found

    msg_id = int(bo_match.group(1))
    sender = bo_match.group(2)

    # Find all SG_ lines after the BO_ line
    remaining_text = dbc_text[bo_match.end():]
    # Stop at next BO_ line
    next_bo = re.search(r"\nBO_\s+", remaining_text)
    if next_bo:
        remaining_text = remaining_text[:next_bo.start()]

    for sg_match in sg_pattern.finditer(remaining_text):
        sig_name = sg_match.group(1)
        start_bit = int(sg_match.group(2))
        bit_length = int(sg_match.group(3))
        endian = int(sg_match.group(4))  # 0=MSB/BE, 1=LSB/LE
        sign_char = sg_match.group(5)
        factor = float(sg_match.group(6))
        offset_val = float(sg_match.group(7))
        min_val = float(sg_match.group(8))
        max_val = float(sg_match.group(9))
        unit = sg_match.group(10)

        is_signed = (sign_char == "-")

        # Determine C type
        if is_signed:
            if bit_length <= 8:
                c_type = "int8_t"
            elif bit_length <= 16:
                c_type = "int16_t"
            elif bit_length <= 32:
                c_type = "int32_t"
            else:
                c_type = "int64_t"
        else:
            if bit_length <= 8:
                c_type = "uint8_t"
            elif bit_length <= 16:
                c_type = "uint16_t"
            elif bit_length <= 32:
                c_type = "uint32_t"
            else:
                c_type = "uint64_t"

        byte_size = 1 if bit_length <= 8 else (2 if bit_length <= 16 else (4 if bit_length <= 32 else 8))

        info = DBCSignalInfo(
            signal_name=sig_name,
            message_name=VIRTUAL_MSG_NAME,
            message_id=msg_id,
            sender=sender,
            is_tx=False,  # Virtual messages don't have real direction
            is_virtual=True,  # Mark as app signal
            c_type=c_type,
            byte_size=byte_size,
            bit_length=bit_length,
            factor=factor,
            offset=offset_val,
            unit=unit,
            min_val=min_val,
            max_val=max_val,
            is_signed=is_signed,
            is_float=False,
        )
        signal_map[sig_name] = info
        print(f"    [Virtual] {sig_name} ({VIRTUAL_MSG_NAME}, {c_type}, {bit_length} bits)")

    return signal_map


def parse_dbc(dbc_path: Path, node_name: str = "SDV_M_CORE0") -> Dict[str, DBCSignalInfo]:
    """
    Parse a DBC file and return a dict mapping signal_name -> DBCSignalInfo.

    For each signal, determine:
    - Whether it belongs to a TX or RX message (relative to node_name)
    - Whether it's a virtual message (VECTOR__INDEPENDENT_SIG_MSG)
    - Its C type and byte size
    """
    db = cantools.database.load_file(str(dbc_path), strict=False)
    signal_map: Dict[str, DBCSignalInfo] = {}

    for msg in db.messages:
        is_virtual = (msg.name == VIRTUAL_MSG_NAME)
        # cantools returns senders as list of strings
        sender_name = msg.senders[0] if isinstance(msg.senders[0], str) else msg.senders[0].name if msg.senders else ""
        is_tx = (sender_name == node_name) if msg.senders else False

        for sig in msg.signals:
            info = DBCSignalInfo(
                signal_name=sig.name,
                message_name=msg.name,
                message_id=msg.frame_id,
                sender=sender_name,
                is_tx=is_tx,
                is_virtual=is_virtual,
                c_type=c_type_for_signal(sig),
                byte_size=byte_size_for_signal(sig),
                bit_length=sig.length,
                factor=sig.scale,
                offset=sig.offset,
                unit=sig.unit or "",
                min_val=sig.minimum if sig.minimum is not None else 0.0,
                max_val=sig.maximum if sig.maximum is not None else 0.0,
                is_signed=sig.is_signed,
                is_float=sig.is_float,
            )
            signal_map[sig.name] = info

    # cantools often skips virtual messages (DLC=0), so parse them manually
    virtual_signals = _parse_dbc_virtual_signals(dbc_path, node_name)
    for sig_name, sig_info in virtual_signals.items():
        if sig_name not in signal_map:
            signal_map[sig_name] = sig_info

    return signal_map


# =====================================================================
#  Excel Parsing
# =====================================================================

def parse_excel(excel_path: Path) -> List[Dict]:
    """
    Parse the S2S_HPC_A_SAG.xlsx workbook and return a list of raw service
    row dicts (skipping empty separator rows).

    Expected columns:
      Service InterFace Name, Service ID, Service Description,
      Method/Event/Field, Setter/Getter/Notifier, Element Name,
      Element Description, Method ID/Event ID, Eventgroup Name,
      Eventgroup ID, Send Strategy, Cyclic Time(ms), Parameter Name,
      IN/OUT, Parameter Description, Parameter Data Type,
      UDP/TCP, Is S2S, Vehicle Define Begin, POC, Vehicle Define End
    """
    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active

    # Read header row to find column indices
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header) if name}

    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for col_name, idx in col_idx.items():
            row_dict[col_name] = row_cells[idx] if idx < len(row_cells) else None

        # Skip empty separator rows (where Service InterFace Name is None/blank)
        svc_name = row_dict.get("Service InterFace Name")
        if svc_name is None or (isinstance(svc_name, str) and svc_name.strip() == ""):
            continue

        rows.append(row_dict)

    wb.close()
    return rows


# =====================================================================
#  Signal Matching Engine
# =====================================================================

def match_signal_to_dbc(param_data_type: str, signal_map: Dict[str, DBCSignalInfo]) -> Optional[DBCSignalInfo]:
    """
    Try to match an Excel 'Parameter Data Type' value to a DBC signal name.

    Multi-level matching strategy:
      1. Exact match: "ParkingSts" → signal "ParkingSts"
      2. Strip suffix "Sts"/"St": "VehicleSpeedSts" → signal "VehicleSpeed"
      3. Add suffix "Sts"/"St": "HighVoltageBattery" → signal "HighVoltageBatterySts"
      4. No match → treat as application-layer signal (not in DBC / virtual)

    IMPORTANT: Prefix containment matching is intentionally NOT used because it
    causes false positives (e.g. "VehicleMode" incorrectly matching "VehicleSpeed").
    """
    # Level 1: Exact match
    if param_data_type in signal_map:
        return signal_map[param_data_type]

    # Level 2: Strip "Sts"/"St" suffix and try again
    for suffix in ["Sts", "St"]:
        if param_data_type.endswith(suffix):
            candidate = param_data_type[:-len(suffix)]
            if candidate in signal_map:
                return signal_map[candidate]

    # Level 3: Add "Sts"/"St" suffix and try again
    for suffix in ["Sts", "St"]:
        candidate = param_data_type + suffix
        if candidate in signal_map:
            return signal_map[candidate]

    # No match — this is an application-layer signal not mapped to DBC
    return None


def parse_service_id(raw) -> int:
    """Parse a Service ID value from Excel (could be "0x0001" or int)."""
    if isinstance(raw, int):
        return raw
    if isinstance(raw, str):
        raw = raw.strip()
        if raw.startswith("0x") or raw.startswith("0X"):
            return int(raw, 16)
        try:
            return int(raw)
        except ValueError:
            return 0
    return 0


def build_service_entries(
    excel_rows: List[Dict],
    signal_map: Dict[str, DBCSignalInfo],
) -> List[SOAServiceEntry]:
    """
    Build the complete list of SOA service entries by:
    1. Processing each Excel row into a SOAServiceEntry
    2. Matching Parameter Data Type to DBC signals
    3. Detecting Setter-Notifier linkages (same ServiceID)
    4. Assigning table indices
    """
    entries: List[SOAServiceEntry] = []

    for row in excel_rows:
        svc_name = row.get("Service InterFace Name", "")
        svc_id_raw = row.get("Service ID", "")
        svc_desc = row.get("Service Description", "")
        svc_type = row.get("Setter/Getter/Notifier", "")
        method_id_raw = row.get("Method ID/Event ID", "")
        eg_name = row.get("Eventgroup Name", "")
        eg_id_raw = row.get("Eventgroup ID", "")
        param_data_type = row.get("Parameter Data Type", "")

        # Skip rows with no service type
        if not svc_type or not isinstance(svc_type, str):
            continue

        # Parse IDs
        service_id = parse_service_id(svc_id_raw)
        method_id = parse_service_id(method_id_raw)
        eventgroup_id = parse_service_id(eg_id_raw) if eg_id_raw else 0

        # Clean up eventgroup name
        if eg_name is None:
            eg_name = ""
        if isinstance(eg_name, float) and math.isnan(eg_name):
            eg_name = ""

        # Clean up description
        if svc_desc is None:
            svc_desc = ""
        if isinstance(svc_desc, float) and math.isnan(svc_desc):
            svc_desc = ""

        # Match signal
        matched_signal = None
        is_app_signal = False
        if param_data_type and isinstance(param_data_type, str):
            matched_signal = match_signal_to_dbc(param_data_type, signal_map)
            if matched_signal is None:
                # No DBC match — this is an application-layer signal.
                # Create a synthetic DBCSignalInfo for it.
                is_app_signal = True
                matched_signal = DBCSignalInfo(
                    signal_name=param_data_type,
                    message_name=VIRTUAL_MSG_NAME,
                    message_id=0,
                    sender="",
                    is_tx=False,
                    is_virtual=True,
                    c_type="uint8_t",  # Default to uint8_t for app signals
                    byte_size=1,
                    bit_length=8,
                    factor=1.0,
                    offset=0.0,
                    unit="",
                    min_val=0.0,
                    max_val=255.0,
                    is_signed=False,
                    is_float=False,
                )
                print(f"  APP_SIGNAL: '{param_data_type}' → application-layer signal "
                      f"(no DBC mapping, using APP_SW_R_W_signal_{param_data_type}())")
            else:
                sig_type = "APP" if matched_signal.is_app_signal else ("TX" if matched_signal.is_can_tx else "RX")
                print(f"  MATCH: '{param_data_type}' -> DBC signal '{matched_signal.signal_name}' "
                      f"({matched_signal.message_name}, {sig_type})")

        entry = SOAServiceEntry(
            service_interface_name=svc_name,
            service_id=service_id,
            service_description=svc_desc,
            service_type=svc_type,
            method_id=method_id,
            eventgroup_name=str(eg_name),
            eventgroup_id=eventgroup_id,
            signal_info=matched_signal,
        )
        entries.append(entry)

    # --- Detect Setter-Notifier linkages (same ServiceID) ---
    # Group by ServiceID
    sid_groups: Dict[int, List[int]] = {}
    for idx, entry in enumerate(entries):
        sid_groups.setdefault(entry.service_id, []).append(idx)

    for sid, indices in sid_groups.items():
        setters = [i for i in indices if entries[i].is_setter]
        notifiers = [i for i in indices if entries[i].is_notifier]
        if setters and notifiers:
            for setter_idx in setters:
                entries[setter_idx].has_linked_notifier = True
                # Link to the first notifier with the same ServiceID
                entries[setter_idx].linked_notifier_idx = notifiers[0]

    # --- Assign table indices ---
    for idx, entry in enumerate(entries):
        entry.table_index = idx

    return entries


# =====================================================================
#  Code Generation — Header (.h)
# =====================================================================

def generate_header(entries: List[SOAServiceEntry], dbc_path: str, excel_path: str) -> str:
    """Generate the SOA_Adapter_Generated.h file content."""
    lines = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    notifier_count = sum(1 for e in entries if e.is_notifier)

    lines.append("/**")
    lines.append(" * @file SOA_Adapter_Generated.h")
    lines.append(" * @brief SOA Adapter Configuration — Auto-Generated Service Matrix Definitions")
    lines.append(" *")
    lines.append(" * Generated by soa_adapter_generator.py")
    lines.append(f" * Date: {now}")
    lines.append(f" * DBC source:  {dbc_path}")
    lines.append(f" * Excel source: {excel_path}")
    lines.append(" *")
    lines.append(" * DO NOT EDIT — re-run the generator to update.")
    lines.append(" */")
    lines.append("")
    lines.append("#ifndef SOA_ADAPTER_GENERATED_H")
    lines.append("#define SOA_ADAPTER_GENERATED_H")
    lines.append("")
    lines.append("#if defined(__cplusplus)")
    lines.append('extern "C"{')
    lines.append("#endif")
    lines.append("")

    # ==========================================================================
    #  IPC-LAYER FIXED ID MAPPING
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                         INCLUDE FILES")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append('#include "Platform.h"')
    lines.append("")

    lines.append("/*==================================================================================================")
    lines.append(" *                                  IPC-LAYER FIXED ID MAPPING")
    lines.append(" *")
    lines.append(" * Per soa.md specification:")
    lines.append(" *   - All Notifier/Event use fixed IPC EventID = 3")
    lines.append(" *   - All Getter/Setter/R-R Method use fixed IPC MethodID = 1")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/** IPC-layer EventID for all SOA Notifier/Event messages (M->A) */")
    lines.append("#define SOA_IPC_EVENT_ID_FOR_NOTIF      (3U)")
    lines.append("")
    lines.append("/** IPC-layer MethodID for all SOA Getter/Setter/Method messages (A->M->A) */")
    lines.append("#define SOA_IPC_METHOD_ID_FOR_RR        (1U)")
    lines.append("")

    # ==========================================================================
    #  SOA PICC CONFIGURATION
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SOA PICC CONFIGURATION")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/** M-Core SOA Provider ID (from ID range 71-80) */")
    lines.append("#define SOA_PROVIDER_ID                 (71U)   /* 0x47 */")
    lines.append("")
    lines.append("/** A-Core SOA Consumer ID (from ID range 71-80) */")
    lines.append("#define SOA_CONSUMER_ID                 (76U)   /* 0x4C */")
    lines.append("")
    lines.append("/** IPCF channel for SOA communication */")
    lines.append("#define SOA_CHANNEL_ID                  (2U)")
    lines.append("")

    # ==========================================================================
    #  SOA PROTOCOL HEADER (12 Bytes)
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SOA PROTOCOL HEADER (12 Bytes)")
    lines.append(" *")
    lines.append(" * Sits inside the IPCF Payload. All fields are Big-Endian (network byte order).")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/** SOA header size in bytes (ServiceID + MethodID + InstanceID + SessionID + ReturnCode + Length) */")
    lines.append("#define SOA_HEADER_SIZE                 (12U)")
    lines.append("")
    lines.append("/** Maximum SOA Payload data size (single message) */")
    lines.append("#define SOA_MAX_DATA_SIZE               (256U)")
    lines.append("")
    lines.append("/** Maximum total SOA message size (header + data) */")
    lines.append("#define SOA_MAX_MSG_SIZE                (SOA_HEADER_SIZE + SOA_MAX_DATA_SIZE)")
    lines.append("")
    lines.append("/**")
    lines.append(" * @brief SOA Protocol Header (12 bytes, Big-Endian)")
    lines.append(" *")
    lines.append(" * Layout inside IPCF Payload:")
    lines.append(" *   [0-1]  SOA_ServiceID")
    lines.append(" *   [2-3]  SOA_MethodID")
    lines.append(" *   [4-5]  SOA_InstanceID")
    lines.append(" *   [6-7]  SOA_SessionID")
    lines.append(" *   [8-9]  SOA_ReturnCode")
    lines.append(" *   [10-11] SOA_Length")
    lines.append(" */")
    lines.append("typedef struct {")
    lines.append("    uint16  SOA_ServiceID;      /**< AP Service ID (from service matrix) */")
    lines.append("    uint16  SOA_MethodID;       /**< AP Method/Event ID (from service matrix) */")
    lines.append("    uint16  SOA_InstanceID;     /**< AP Instance ID (from service matrix) */")
    lines.append("    uint16  SOA_SessionID;      /**< Session ID (0 for Notifier/Event, echoed for Getter/Setter) */")
    lines.append("    uint16  SOA_ReturnCode;     /**< Return code: 0=success, non-0=failure */")
    lines.append("    uint16  SOA_Length;         /**< Length of actual parameter data following this header */")
    lines.append("} SOA_Header_t;")
    lines.append("")

    # ==========================================================================
    #  SOA SERVICE TYPE ENUMERATION
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SOA SERVICE TYPE ENUMERATION")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/**")
    lines.append(" * @brief SOA service type")
    lines.append(" */")
    lines.append("typedef enum {")
    lines.append("    SOA_SERVICE_NOTIFIER = 0U,  /**< Field Notifier (M->A, Event, update-on-changed) */")
    lines.append("    SOA_SERVICE_GETTER   = 1U,  /**< Field Getter (A requests, M responds with current value) */")
    lines.append("    SOA_SERVICE_SETTER   = 2U   /**< Field Setter (A requests write, M writes and responds) */")
    lines.append("} SOA_ServiceType_e;")
    lines.append("")

    # ==========================================================================
    #  SIGNAL READ/WRITE FUNCTION TYPES
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SIGNAL READ/WRITE FUNCTION TYPES")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/**")
    lines.append(" * @brief Read a signal value into a buffer")
    lines.append(" *")
    lines.append(" * @param[out] outBuf   Buffer to write the signal value into (serialized, big-endian)")
    lines.append(" * @param[in]  maxLen   Maximum buffer size")
    lines.append(" * @return Actual number of bytes written, 0 on error")
    lines.append(" */")
    lines.append("typedef uint16 (*SOA_SignalReadFunc_t)(uint8 *outBuf, uint16 maxLen);")
    lines.append("")
    lines.append("/**")
    lines.append(" * @brief Write a signal value from a buffer")
    lines.append(" *")
    lines.append(" * @param[in]  inBuf    Buffer containing the value to write (big-endian)")
    lines.append(" * @param[in]  len      Number of bytes in the buffer")
    lines.append(" * @return 0 on success, non-zero on error")
    lines.append(" */")
    lines.append("typedef uint8 (*SOA_SignalWriteFunc_t)(const uint8 *inBuf, uint16 len);")
    lines.append("")

    # ==========================================================================
    #  SOA SERVICE CONFIGURATION TABLE STRUCT
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SOA SERVICE CONFIGURATION TABLE")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/**")
    lines.append(" * @brief SOA service configuration entry")
    lines.append(" *")
    lines.append(" * One entry per service defined in the SOA service matrix.")
    lines.append(" * The table is instantiated in SOA_Adapter_Generated.c.")
    lines.append(" */")
    lines.append("typedef struct {")
    lines.append("    uint16              SOA_ServiceID;      /**< AP Service ID */")
    lines.append("    uint16              SOA_MethodID;       /**< AP Method/Event ID */")
    lines.append("    uint16              SOA_InstanceID;     /**< AP Instance ID */")
    lines.append("    SOA_ServiceType_e   serviceType;        /**< Notifier / Getter / Setter */")
    lines.append("    SOA_SignalReadFunc_t  readFunc;         /**< Read signal value (for Notifier and Getter) */")
    lines.append("    SOA_SignalWriteFunc_t writeFunc;        /**< Write signal value (for Setter only, NULL otherwise) */")
    lines.append("    uint16              SOA_EventGroupID;   /**< Eventgroup ID (Notifier only, 0 for others) */")
    lines.append("    uint8               dataSize;           /**< Expected data size in bytes for this signal */")
    lines.append("    boolean             hasLinkedNotifier;  /**< TRUE if this Setter has a linked Notifier */")
    lines.append("    uint8               linkedNotifierIdx;  /**< Index into service table of the linked Notifier */")
    lines.append("} SOA_ServiceConfig_t;")
    lines.append("")

    # ==========================================================================
    #  SERVICE MATRIX DEFINITIONS
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SERVICE MATRIX DEFINITIONS")
    lines.append(" *")

    # Build a summary table
    lines.append(" * From the SOA service matrix:")
    lines.append(" *   | Idx | Service Interface             | ServiceID | Type     | SOA MethodID | Signal           |")
    lines.append(" *   |-----|-------------------------------|-----------|----------|-------------|------------------|")
    for e in entries:
        sig_name = e.signal_info.signal_name if e.signal_info else "(none)"
        lines.append(f" *   | {e.table_index:3d} | {e.service_interface_name:<29s} | {format_hex16(e.service_id):>8s} | {e.service_type:<8s} | {format_hex16(e.method_id):>11s} | {sig_name:<16s} |")

    lines.append(" *==================================================================================================*/")
    lines.append("")

    # --- Service table indices ---
    lines.append("/* --- Service table indices --- */")
    for e in entries:
        lines.append(f"#define {e.idx_macro_name:<45s} ({e.table_index}U)")
    lines.append("")

    # --- Totals ---
    lines.append(f"/** Total number of SOA services */")
    lines.append(f"#define SOA_SERVICE_TABLE_COUNT         ({len(entries)}U)")
    lines.append("")
    lines.append(f"/** Number of Notifier services (for initial value sync and change detection) */")
    lines.append(f"#define SOA_NOTIFIER_COUNT              ({notifier_count}U)")
    lines.append("")

    # --- ServiceID definitions ---
    lines.append("/* --- ServiceID definitions --- */")
    seen_sids = set()
    for e in entries:
        if e.service_id not in seen_sids:
            lines.append(f"#define {e.sid_macro_name:<45s} ({format_hex16(e.service_id)})")
            seen_sids.add(e.service_id)
    lines.append("")

    # --- MethodID/EventID definitions ---
    lines.append("/* --- SOA MethodID / EventID definitions --- */")
    for e in entries:
        comment = f"/**< {e.service_type}: {e.signal_info.signal_name if e.signal_info else e.short_name} */"
        lines.append(f"#define {e.mid_macro_name:<45s} ({format_hex16(e.method_id)})  {comment}")
    lines.append("")

    # --- InstanceID ---
    lines.append("/* --- InstanceID (all services use instance 1) --- */")
    lines.append("#define SOA_INSTANCE_ID_DEFAULT         (0x0001U)")
    lines.append("")

    # --- EventGroup IDs ---
    eg_macros = {}
    for e in entries:
        if e.eg_macro_name and e.eg_macro_name not in eg_macros:
            eg_macros[e.eg_macro_name] = e.eventgroup_id
    if eg_macros:
        lines.append("/* --- EventGroup IDs --- */")
        for eg_name, eg_val in eg_macros.items():
            lines.append(f"#define {eg_name:<45s} ({format_hex16(eg_val)})")
        lines.append("")

    # ==========================================================================
    #  GLOBAL DECLARATIONS
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  GLOBAL DECLARATIONS")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/** SOA service configuration table (defined in SOA_Adapter_Generated.c) */")
    lines.append("extern const SOA_ServiceConfig_t g_soaServiceTable[SOA_SERVICE_TABLE_COUNT];")
    lines.append("")
    lines.append("/** Notifier index mapping (indices into g_soaServiceTable for Notifier-type services) */")
    lines.append("extern const uint8 g_soaNotifierIndices[SOA_NOTIFIER_COUNT];")
    lines.append("")

    # ==========================================================================
    #  SIGNAL READ/WRITE FUNCTION DECLARATIONS
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                                  SIGNAL READ/WRITE FUNCTION DECLARATIONS")
    lines.append(" *==================================================================================================*/")
    lines.append("")

    # Notifier read functions
    notifier_entries = [e for e in entries if e.is_notifier]
    if notifier_entries:
        lines.append("/* Notifier read functions */")
        for e in notifier_entries:
            lines.append(f"uint16 {e.read_func_name}(uint8 *outBuf, uint16 maxLen);")
        lines.append("")

    # Getter read functions
    getter_entries = [e for e in entries if e.is_getter]
    if getter_entries:
        lines.append("/* Getter read functions */")
        for e in getter_entries:
            lines.append(f"uint16 {e.read_func_name}(uint8 *outBuf, uint16 maxLen);")
        lines.append("")

    # Setter write functions
    setter_entries = [e for e in entries if e.is_setter]
    if setter_entries:
        lines.append("/* Setter write functions */")
        for e in setter_entries:
            lines.append(f"uint8  {e.write_func_name}(const uint8 *inBuf, uint16 len);")
        lines.append("")

    lines.append("#if defined(__cplusplus)")
    lines.append("}")
    lines.append("#endif")
    lines.append("")
    lines.append("#endif /* SOA_ADAPTER_GENERATED_H */")

    return "\n".join(lines) + "\n"


# =====================================================================
#  Code Generation — Source (.c)
# =====================================================================

def generate_source(entries: List[SOAServiceEntry], dbc_path: str, excel_path: str) -> str:
    """Generate the SOA_Adapter_Generated.c file content."""
    lines = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines.append("/**")
    lines.append(" * @file SOA_Adapter_Generated.c")
    lines.append(" * @brief SOA Adapter Configuration — Auto-Generated Signal Read/Write Functions and Service Table")
    lines.append(" *")
    lines.append(" * Generated by soa_adapter_generator.py")
    lines.append(f" * Date: {now}")
    lines.append(f" * DBC source:  {dbc_path}")
    lines.append(f" * Excel source: {excel_path}")
    lines.append(" *")
    lines.append(" * Signal Mapping:")

    for e in entries:
        if e.signal_info:
            sig = e.signal_info
            direction = "M→A" if e.is_notifier else ("A→M→A" if e.is_getter else "A→M")
            access = sig.can_access_func
            if sig.is_app_signal:
                sig_type = "APP"
            elif sig.is_can_rx:
                sig_type = "RX"
            else:
                sig_type = "TX"
            lines.append(f" *   {e.service_type:<8s} {sig.signal_name:<24s} → {access}() ({sig_type} {sig.message_name})")
        else:
            lines.append(f" *   {e.service_type:<8s} {e.short_name:<24s} → (no DBC signal matched)")
    lines.append(" */")
    lines.append("")
    lines.append("#if defined(__cplusplus)")
    lines.append('extern "C"{')
    lines.append("#endif")
    lines.append("")
    lines.append("/*==================================================================================================")
    lines.append(" *                                         INCLUDE FILES")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append('#include "SOA_Adapter_Generated.h"')
    lines.append('#include "CANdbc_Generated.h"   /* DBC global structs and signal access functions */')
    lines.append("")

    # ==========================================================================
    #  SIGNAL READ FUNCTIONS (Notifier + Getter)
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                              SIGNAL READ FUNCTIONS (Notifier + Getter)")
    lines.append(" *")
    lines.append(" * Each function reads the current value from the DBC global variable,")
    lines.append(" * serializes it into big-endian byte order, and returns the byte count.")
    lines.append(" *==================================================================================================*/")
    lines.append("")

    for e in entries:
        if not e.has_read_func:
            continue

        sig = e.signal_info
        if sig is None:
            # No matched signal — generate a placeholder
            lines.append(f"/*")
            lines.append(f" * @brief Read {e.short_name} signal ({e.service_type}: {e.service_interface_name})")
            lines.append(f" *")
            lines.append(f" * WARNING: No DBC signal matched. Placeholder returns 0 bytes.")
            lines.append(f" */")
            lines.append(f"uint16 {e.read_func_name}(uint8 *outBuf, uint16 maxLen)")
            lines.append(f"{{")
            lines.append(f"    (void)outBuf;")
            lines.append(f"    (void)maxLen;")
            lines.append(f"    return 0U;")
            lines.append(f"}}")
            lines.append("")
            continue

        byte_size = sig.byte_size
        c_type = c_type_for_byte_size(byte_size)

        lines.append(f"/**")
        lines.append(f" * @brief Read {sig.signal_name} signal ({e.service_type}: {e.service_interface_name})")
        lines.append(f" *")

        if sig.is_can_rx:
            lines.append(f" * Source: {sig.message_name} (RX, sender={sig.sender})")
        elif sig.is_can_tx:
            lines.append(f" * Source: {sig.message_name} (TX, sender={sig.sender})")
        else:
            lines.append(f" * Source: {sig.message_name} (virtual/app signal)")
        lines.append(f" * Serialized as {byte_size} byte{'s' if byte_size > 1 else ''}, big-endian.")
        lines.append(f" */")
        lines.append(f"uint16 {e.read_func_name}(uint8 *outBuf, uint16 maxLen)")
        lines.append(f"{{")
        lines.append(f"    if ((outBuf == NULL_PTR) || (maxLen < {byte_size}U))")
        lines.append(f"    {{")
        lines.append(f"        return 0U;")
        lines.append(f"    }}")

        # Read from signal
        if sig.is_can_rx:
            lines.append(f"    {c_type} val = {sig.can_access_func}();")
        else:
            # TX or app signal: use R_W with NULL
            lines.append(f"    {c_type} val = {sig.can_access_func}(NULL);")

        # Serialize big-endian
        if byte_size == 1:
            lines.append(f"    outBuf[0] = (uint8)(val & 0xFFU);")
        elif byte_size == 2:
            lines.append(f"    outBuf[0] = (uint8)(val >> 8U);")
            lines.append(f"    outBuf[1] = (uint8)(val & 0xFFU);")
        elif byte_size == 4:
            lines.append(f"    outBuf[0] = (uint8)(val >> 24U);")
            lines.append(f"    outBuf[1] = (uint8)(val >> 16U);")
            lines.append(f"    outBuf[2] = (uint8)(val >> 8U);")
            lines.append(f"    outBuf[3] = (uint8)(val & 0xFFU);")
        elif byte_size == 8:
            lines.append(f"    outBuf[0] = (uint8)(val >> 56U);")
            lines.append(f"    outBuf[1] = (uint8)(val >> 48U);")
            lines.append(f"    outBuf[2] = (uint8)(val >> 40U);")
            lines.append(f"    outBuf[3] = (uint8)(val >> 32U);")
            lines.append(f"    outBuf[4] = (uint8)(val >> 24U);")
            lines.append(f"    outBuf[5] = (uint8)(val >> 16U);")
            lines.append(f"    outBuf[6] = (uint8)(val >> 8U);")
            lines.append(f"    outBuf[7] = (uint8)(val & 0xFFU);")

        lines.append(f"    return {byte_size}U;")
        lines.append(f"}}")
        lines.append("")

    # ==========================================================================
    #  SIGNAL WRITE FUNCTIONS (Setter)
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                              SIGNAL WRITE FUNCTIONS (Setter)")
    lines.append(" *==================================================================================================*/")
    lines.append("")

    for e in entries:
        if not e.has_write_func:
            continue

        sig = e.signal_info
        if sig is None:
            # No matched signal — generate a placeholder
            lines.append(f"/*")
            lines.append(f" * @brief Write {e.short_name} signal ({e.service_type}: {e.service_interface_name})")
            lines.append(f" *")
            lines.append(f" * WARNING: No DBC signal matched. Placeholder returns error.")
            lines.append(f" */")
            lines.append(f"uint8 {e.write_func_name}(const uint8 *inBuf, uint16 len)")
            lines.append(f"{{")
            lines.append(f"    (void)inBuf;")
            lines.append(f"    (void)len;")
            lines.append(f"    return 1U;  /* Error */")
            lines.append(f"}}")
            lines.append("")
            continue

        byte_size = sig.byte_size
        c_type = c_type_for_byte_size(byte_size)

        lines.append(f"/**")
        lines.append(f" * @brief Write {sig.signal_name} signal ({e.service_type}: {e.service_interface_name})")
        lines.append(f" *")

        if sig.is_app_signal:
            lines.append(f" * Target: {sig.message_name} (app signal)")
        elif sig.is_can_tx:
            lines.append(f" * Target: {sig.message_name} (TX, sender={sig.sender})")
        else:
            lines.append(f" * Target: {sig.message_name} (RX signal — Setter on RX is unusual)")

        lines.append(f" * Expects {byte_size} byte{'s' if byte_size > 1 else ''} input.")
        lines.append(f" */")
        lines.append(f"uint8 {e.write_func_name}(const uint8 *inBuf, uint16 len)")
        lines.append(f"{{")
        lines.append(f"    if ((inBuf == NULL_PTR) || (len < {byte_size}U))")
        lines.append(f"    {{")
        lines.append(f"        return 1U;  /* Error */")
        lines.append(f"    }}")

        # Deserialize from big-endian
        if byte_size == 1:
            lines.append(f"    {c_type} val = inBuf[0];")
        elif byte_size == 2:
            lines.append(f"    {c_type} val = (({c_type})inBuf[0] << 8U) | ({c_type})inBuf[1];")
        elif byte_size == 4:
            lines.append(f"    {c_type} val = (({c_type})inBuf[0] << 24U) | (({c_type})inBuf[1] << 16U) | (({c_type})inBuf[2] << 8U) | ({c_type})inBuf[3];")
        elif byte_size == 8:
            lines.append(f"    {c_type} val = (({c_type})inBuf[0] << 56U) | (({c_type})inBuf[1] << 48U) | (({c_type})inBuf[2] << 40U) | (({c_type})inBuf[3] << 32U) | (({c_type})inBuf[4] << 24U) | (({c_type})inBuf[5] << 16U) | (({c_type})inBuf[6] << 8U) | ({c_type})inBuf[7];")

        # Write via access function
        lines.append(f"    {sig.can_access_func}(&val);")
        lines.append(f"")
        lines.append(f"    return 0U;  /* Success */")
        lines.append(f"}}")
        lines.append("")

    # ==========================================================================
    #  SOA SERVICE CONFIGURATION TABLE
    # ==========================================================================
    lines.append("/*==================================================================================================")
    lines.append(" *                              SOA SERVICE CONFIGURATION TABLE")
    lines.append(" *==================================================================================================*/")
    lines.append("")
    lines.append("/**")
    lines.append(" * @brief SOA service configuration table")
    lines.append(" *")
    lines.append(" * Index order matches SOA_SVC_IDX_xxx defines in SOA_Adapter_Generated.h.")
    lines.append(" */")
    lines.append(f"const SOA_ServiceConfig_t g_soaServiceTable[SOA_SERVICE_TABLE_COUNT] =")
    lines.append(f"{{")

    for e in entries:
        sig = e.signal_info
        sig_label = sig.signal_name if sig else e.short_name

        lines.append(f"    /* [{e.table_index}] {e.service_interface_name} — {e.service_type}: {sig_label} */")
        lines.append(f"    {{")

        # ServiceID
        sid_macro = e.sid_macro_name
        lines.append(f"        .SOA_ServiceID      = {sid_macro},       /* {format_hex16(e.service_id)} */")

        # MethodID
        mid_macro = e.mid_macro_name
        lines.append(f"        .SOA_MethodID       = {mid_macro},     /* {format_hex16(e.method_id)} */")

        # InstanceID
        lines.append(f"        .SOA_InstanceID     = SOA_INSTANCE_ID_DEFAULT,      /* 0x0001 */")

        # serviceType
        lines.append(f"        .serviceType        = {e.soa_service_type_enum},")

        # readFunc
        if e.has_read_func:
            lines.append(f"        .readFunc           = {e.read_func_name},")
        else:
            lines.append(f"        .readFunc           = NULL_PTR,                      /* Setter: no readFunc, uses linked Notifier */")

        # writeFunc
        if e.has_write_func:
            lines.append(f"        .writeFunc          = {e.write_func_name},")
        else:
            lines.append(f"        .writeFunc          = NULL_PTR,")

        # EventGroupID
        if e.is_notifier and e.eg_macro_name:
            lines.append(f"        .SOA_EventGroupID   = {e.eg_macro_name},       /* {format_hex16(e.eventgroup_id)} */")
        else:
            lines.append(f"        .SOA_EventGroupID   = 0U,")

        # dataSize
        lines.append(f"        .dataSize           = {e.data_size}U,")

        # hasLinkedNotifier
        if e.has_linked_notifier:
            lines.append(f"        .hasLinkedNotifier  = TRUE,")
        else:
            lines.append(f"        .hasLinkedNotifier  = FALSE,")

        # linkedNotifierIdx
        if e.has_linked_notifier:
            notifier_entry = entries[e.linked_notifier_idx]
            lines.append(f"        .linkedNotifierIdx  = {notifier_entry.idx_macro_name}  /* Index {e.linked_notifier_idx} */")
        else:
            lines.append(f"        .linkedNotifierIdx  = 0U")

        lines.append(f"    }},")
        lines.append(f"")

    lines.append(f"}};")
    lines.append(f"")

    # ==========================================================================
    #  NOTIFIER INDEX ARRAY
    # ==========================================================================
    notifier_entries = [e for e in entries if e.is_notifier]

    lines.append("/**")
    lines.append(" * @brief Notifier index mapping")
    lines.append(" *")
    lines.append(" * Lists indices into g_soaServiceTable for all Notifier-type services.")
    lines.append(" * Used for initial value sync and change detection.")
    lines.append(" */")
    lines.append(f"const uint8 g_soaNotifierIndices[SOA_NOTIFIER_COUNT] =")
    lines.append(f"{{")

    for i, e in enumerate(notifier_entries):
        comma = "," if i < len(notifier_entries) - 1 else ""
        lines.append(f"    {e.idx_macro_name}{comma}    /* [{i}] → g_soaServiceTable[{e.table_index}] */")

    lines.append(f"}};")
    lines.append(f"")

    lines.append("#if defined(__cplusplus)")
    lines.append("}")
    lines.append("#endif")

    return "\n".join(lines) + "\n"


# =====================================================================
#  Main
# =====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="SOA Adapter Code Generator — reads DBC + Excel, generates SOA_Adapter_Generated.c/.h"
    )
    parser.add_argument(
        "--dbc", type=Path, default=DEFAULT_DBC_FILE_PATH,
        help="Path to the input .dbc file"
    )
    parser.add_argument(
        "--excel", type=Path, default=DEFAULT_EXCEL_FILE_PATH,
        help="Path to the input S2S_HPC_A_SAG.xlsx file"
    )
    parser.add_argument(
        "-o", "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR,
        help="Directory for generated output files"
    )
    parser.add_argument(
        "--node", type=str, default="SDV_M_CORE0",
        help="M-Core node name in DBC (default: SDV_M_CORE0)"
    )
    parser.add_argument(
        "--basename", type=str, default=DEFAULT_OUTPUT_BASENAME,
        help="Base name for generated files (default: SOA_Adapter_Generated)"
    )
    args = parser.parse_args()

    dbc_path = args.dbc.resolve()
    excel_path = args.excel.resolve()
    output_dir = args.output_dir.resolve()

    # Validate inputs
    if not dbc_path.is_file():
        print(f"ERROR: DBC file not found: {dbc_path}", file=sys.stderr)
        sys.exit(1)
    if not excel_path.is_file():
        print(f"ERROR: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"=" * 60)
    print(f"SOA Adapter Code Generator")
    print(f"=" * 60)
    print(f"  DBC:    {dbc_path}")
    print(f"  Excel:  {excel_path}")
    print(f"  Output: {output_dir}")
    print(f"  Node:   {args.node}")
    print()

    # --- Step 1: Parse DBC ---
    print("[1/4] Parsing DBC file...")
    signal_map = parse_dbc(dbc_path, node_name=args.node)
    print(f"  Found {len(signal_map)} signals across all messages:")
    for sig_name, sig_info in signal_map.items():
        sig_type = "APP" if sig_info.is_app_signal else ("TX" if sig_info.is_can_tx else "RX")
        print(f"    {sig_name:<24s} ({sig_info.message_name}, {sig_type}, {sig_info.c_type})")
    print()

    # --- Step 2: Parse Excel ---
    print("[2/4] Parsing Excel service matrix...")
    excel_rows = parse_excel(excel_path)
    print(f"  Found {len(excel_rows)} service entries:")
    for row in excel_rows:
        svc_name = row.get("Service InterFace Name", "?")
        svc_type = row.get("Setter/Getter/Notifier", "?")
        param_dt = row.get("Parameter Data Type", "?")
        print(f"    {svc_name:<30s}  {svc_type:<10s}  ParamType={param_dt}")
    print()

    # --- Step 3: Build service entries with signal matching ---
    print("[3/4] Building service entries with signal matching...")
    entries = build_service_entries(excel_rows, signal_map)
    print(f"  Built {len(entries)} service entries:")
    for e in entries:
        sig_str = f"→ {e.signal_info.signal_name} ({e.signal_info.message_name})" if e.signal_info else "(no match)"
        link_str = f" [linked→idx {e.linked_notifier_idx}]" if e.has_linked_notifier else ""
        print(f"    [{e.table_index}] {e.service_interface_name:<30s} {e.service_type:<8s} {sig_str}{link_str}")
    print()

    # --- Step 4: Generate C files ---
    print("[4/4] Generating C code...")

    header_path = output_dir / f"{args.basename}.h"
    source_path = output_dir / f"{args.basename}.c"

    header_content = generate_header(entries, str(dbc_path), str(excel_path))
    source_content = generate_source(entries, str(dbc_path), str(excel_path))

    header_path.write_text(header_content, encoding="utf-8")
    source_path.write_text(source_content, encoding="utf-8")

    print(f"  Generated: {header_path}")
    print(f"  Generated: {source_path}")
    print()
    print(f"Done! {len(entries)} services, {sum(1 for e in entries if e.is_notifier)} notifiers.")


if __name__ == "__main__":
    main()
